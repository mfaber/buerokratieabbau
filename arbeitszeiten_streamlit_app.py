import streamlit as st
import pandas as pd
import pdfplumber
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# ---- Streamlit UI ----
st.set_page_config(page_title="Arbeitszeiten-Extraktion", layout="wide")
st.title("🕒 Arbeitszeit-Extraktion aus MyTMA-PDF")

st.markdown("""## ℹ️ Anleitung

1. **Zeiten aus MyTMA exportieren:** Auskunft → **Selbstauskunft**, Monat/Jahr wählen, Haken bei **„Bemerkungen“** und **„Kalenderwochen“** entfernen, dann **Drucken** → PDF speichern.
2. **PDF hochladen**, das aus MyTMA exportiert wurde.
3. Das Tool extrahiert je Tag das **früheste „Von“** und das **späteste „Bis“** (über alle Buchungen des Tages). Dadurch funktionieren auch Tage mit **mehr als zwei Intervallen**.
4. Ergebnis prüfen und als **Excel** herunterladen.
""")

uploaded_file = st.file_uploader("PDF-Datei hochladen", type="pdf")

# ---- Parsing helpers ----

def _parse_header_month(text: str):
    """Liest den Zielmonat aus 'Zeitraum: dd.mm.yyyy Bis dd.mm.yyyy'."""
    m = re.search(r'Zeitraum:\s*(\d{2}\.\d{2}\.\d{4})\s*Bis\s*(\d{2}\.\d{2}\.\d{4})', text or "")
    if not m:
        return None, None
    end = m.group(2)
    return end[3:5], end[6:10]  # month, year

def extract_times_from_pdf(pdf_bytes: bytes) -> pd.DataFrame:
    """Robuste Extraktion mittels **Spaltenpositionen**:
    - Ermittelt die x-Position der Überschriften *Von/Bis/Von/Bis* auf jeder Seite.
    - Sammelt für jede Tageszeile **nur** die Zeiten, die in diese Spalten fallen.
    - Bestimmt daraus *Von_gesamt* (Minimum aller 'Von') und *Bis_gesamt* (Maximum aller 'Bis').
    Fallbacks für leere Seiten/abweichende Layouts sind enthalten.
    """
    # Zielmonat aus Seite 1 bestimmen (zum Filtern von Vormonats-/Folge-Tagen)
    target_month, target_year = None, None
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        if pdf.pages:
            header_text = pdf.pages[0].extract_text(x_tolerance=2) or ""
            target_month, target_year = _parse_header_month(header_text)

    records = []

    def to_min(t: str) -> int:
        h, m = map(int, t.split(':'))
        return h * 60 + m

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words = page.extract_words(
                x_tolerance=1,
                y_tolerance=3,
                keep_blank_chars=False,
                use_text_flow=True,
            )
            # Überschriften auf der Seite finden
            hdr_idx = next((i for i, w in enumerate(words) if w['text'] == 'Datum'), None)
            if hdr_idx is None:
                continue
            hdr = words[hdr_idx:hdr_idx + 12]
            try:
                tm_x   = hdr[2]['x0']
                von1_x = hdr[3]['x0']
                bis1_x = hdr[4]['x0']
                von2_x = hdr[5]['x0']
                bis2_x = hdr[6]['x0']
                brutto_x = hdr[7]['x0']
            except Exception:
                continue

            # Spalten-Grenzen als Mittelwerte zwischen den Überschriften
            boundaries = [
                ((tm_x + von1_x) / 2, (von1_x + bis1_x) / 2, 'von'),
                ((von1_x + bis1_x) / 2, (bis1_x + von2_x) / 2, 'bis'),
                ((bis1_x + von2_x) / 2, (von2_x + bis2_x) / 2, 'von'),
                ((von2_x + bis2_x) / 2, (bis2_x + brutto_x) / 2, 'bis'),
            ]

            current = None
            for w in words[hdr_idx + 10:]:
                t = w['text']

                # Neue Tageszeile beginnt mit 'dd.mm.'
                if re.match(r'^\d{2}\.\d{2}\.$', t):
                    if (target_month is None) or (t[3:5] == target_month):
                        current = {'Datum': t, 'Wochentag': '', 'von': [], 'bis': []}
                        records.append(current)
                    else:
                        current = None
                    continue

                if current is None:
                    continue

                if not current['Wochentag'] and re.match(r'^[A-Za-zÄÖÜäöü]{2}$', t):
                    current['Wochentag'] = t
                    continue

                if re.match(r'^\d{1,2}:\d{2}$', t):
                    x = w['x0']
                    for xlo, xhi, kind in boundaries:
                        if x >= xlo and x < xhi:
                            current[kind].append(t)
                            break

    # In DataFrame umwandeln
    rows = []
    for ent in records:
        vons = ent['von']
        biss = ent['bis']

        von_gesamt = min(vons, key=to_min) if vons else (min(biss, key=to_min) if biss else '')
        bis_gesamt = max(biss, key=to_min) if biss else (max(vons, key=to_min) if vons else '')

        row = {
            'Datum': ent['Datum'],
            'Wochentag': ent['Wochentag'],
            'Von1': vons[0] if len(vons) > 0 else '',
            'Bis1': biss[0] if len(biss) > 0 else '',
            'Von2': vons[1] if len(vons) > 1 else '',
            'Bis2': biss[1] if len(biss) > 1 else '',
            'Von_gesamt': von_gesamt,
            'Bis_gesamt': bis_gesamt,
        }
        rows.append(row)

    df = pd.DataFrame(rows, columns=['Datum','Wochentag','Von1','Bis1','Von2','Bis2','Von_gesamt','Bis_gesamt'])

    # Zeiten in Stunden/Minuten zerlegen und validieren (nur für *_gesamt streng)
    def parse_time(text):
        m = re.match(r'(\d{1,2})[:\.]?(\d{2})', str(text))
        if m:
            return int(m.group(1)), int(m.group(2))
        return pd.NA, pd.NA

    for col in ['Von1','Bis1','Von2','Bis2','Von_gesamt','Bis_gesamt']:
        df[f"{col}_Stunde"], df[f"{col}_Minute"] = zip(*df[col].apply(parse_time))

    def valid(st, mi):
        return pd.notna(st) and 0 <= st <= 23 and 0 <= mi <= 59

    for col in ['Von_gesamt', 'Bis_gesamt']:
        stc, mic = f"{col}_Stunde", f"{col}_Minute"
        mask = ~df[[stc, mic]].apply(lambda x: valid(x.iloc[0], x.iloc[1]), axis=1)
        df.loc[mask, [col, stc, mic]] = pd.NA

    df = df.astype({c: 'Int64' for c in df.columns if c.endswith('_Stunde') or c.endswith('_Minute')})
    return df

# ---- Excel Export ----

def create_formatted_excel(df: pd.DataFrame) -> bytes:
    blue_border_thin = Side(style='thin', color='0000FF')
    blue_border_thick = Side(style='medium', color='0000FF')
    wb = Workbook()
    ws = wb.active

    # Überschriften
    ws['A1'] = 'Datum'
    ws['B1'] = 'Wochentag'
    ws['D1'] = 'Beginn'
    ws['F1'] = 'Ende'
    ws.merge_cells('D1:E1')
    ws.merge_cells('F1:G1')

    ws['D2'] = 'Std'
    ws['E2'] = 'Min'
    ws['F2'] = 'Std'
    ws['G2'] = 'Min'

    # Spaltenbreiten
    for col, width in zip('ABCDEFG', [10, 10, 2, 6, 5, 6, 5]):
        ws.column_dimensions[col].width = width

    # Daten ab Zeile 3
    for i, row in df.iterrows():
        r = i + 3
        ws.cell(row=r, column=1).value = row['Datum']
        ws.cell(row=r, column=2).value = row['Wochentag']
        ws.cell(row=r, column=4).value = int(row['Von_gesamt_Stunde']) if pd.notna(row['Von_gesamt_Stunde']) else None
        ws.cell(row=r, column=5).value = int(row['Von_gesamt_Minute']) if pd.notna(row['Von_gesamt_Minute']) else None
        ws.cell(row=r, column=6).value = int(row['Bis_gesamt_Stunde']) if pd.notna(row['Bis_gesamt_Stunde']) else None
        ws.cell(row=r, column=7).value = int(row['Bis_gesamt_Minute']) if pd.notna(row['Bis_gesamt_Minute']) else None

    # Rahmen & Wochenenden
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
        wtag = row[1].value  # B-Spalte
        is_weekend = str(wtag).strip() in {'Sa', 'So'}
        for cell in row:
            cell.border = Border(top=blue_border_thin, bottom=blue_border_thin,
                                 left=blue_border_thin, right=blue_border_thin)
            if is_weekend:
                cell.fill = PatternFill('solid', fgColor='FFFF99')
        # dicke Rahmen um Beginn/Ende-Blöcke
        for idx in [3,4,5,6]:  # C..F  (sichtbar sind D..G, aber Rahmen für Block)
            row[idx].border = Border(top=blue_border_thick, bottom=blue_border_thick,
                                     left=blue_border_thick, right=blue_border_thick)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ---- Run ----

if uploaded_file:
    pdf_bytes = uploaded_file.read()
    with st.spinner("Verarbeite PDF..."):
        df_result = extract_times_from_pdf(pdf_bytes)

    st.success("Extraktion abgeschlossen!")

    # Ansicht ohne Hilfsspalten
    drop_cols = [c for c in df_result.columns if c.endswith('_Stunde') or c.endswith('_Minute')]
    df_view = df_result.drop(columns=drop_cols, errors='ignore')
    st.dataframe(df_view, use_container_width=True)

    excel_bytes = create_formatted_excel(df_result)
    st.download_button(
        "📥 Excel herunterladen",
        excel_bytes,
        file_name='Arbeitszeiten_Export_formatiert.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
